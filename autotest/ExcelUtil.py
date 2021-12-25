# coding=utf-8

import sys
import traceback
import openpyxl




"""
 * @ClassName:  ExcelUtil   
 * @Description:python基于openpyxl操作excel
 * @author: xingchunyu
 * @date:   xxxx年xx月xx日 上午/下午xx:xx:xx    
"""
class ExcelUtil:

    def __init__(self, fileName):
        self.fileName = fileName    #初始化Excel文件路

    #新建excel
    def createWb1(self, fileName):
        #初始化
        wb = None
        try:
            wb = openpyxl.Workbook(fileName)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return wb



    #获取workbook类对象
    def getWb1(self, fileName):
        #初始化
        wb = None
        try:
            wb = openpyxl.load_workbook(fileName)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return wb


    #新建sheet(指定新建sheet名及位置)
    def createSheet1(self, wb, sheet_name, index):
        #初始化
        sh = None
        try:
            sh = wb.create_sheet(sheet_name, index)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return sh


    #新建sheet(默认新增在最后位置)
    def createSheet2(self, wb, sheet_name):
        #初始化
        sh = None
        try:
            sh = wb.create_sheet(sheet_name)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return sh


    #删除sheet(指定sheet对象)
    def removeSheet1(self, wb, sh):
        #初始化
        try:
            wb.remove_sheet(sh)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass


    #获取sheet类对象-根据sheet位置（已经存在的excel）
    def getSheet1(self, wb, index):
        #初始化
        sheets = None
        sh = None
        try:
            #获取全部sheet名字
            sheets = wb.sheetnames
            sh = wb[sheets[index]]
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return sh


    #获取sheet类对象-根据sheet名字（已经存在的excel）
    def getSheet2(self, wb, sheet_name):
        #初始化
        sh = None
        try:
            sh = wb[sheet_name]
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return sh


    #拷贝sheet(指定sheet对象)
    def copeSheet1(self, wb, sh):
        #初始化
        try:
            wb.copy_worksheet(from_worksheet=sh)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass


    #修改sheet名(指定sheet对象)
    def modifySheet_name1(self, sh, sheet_new_name):
        #初始化
        try:
            sh.title = sheet_new_name
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass



    #获取sheet名(指定sheet对象)
    def getSheet_name1(self, sh):
        #初始化
        try:
            sheet_name = sh.title
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return sheet_name


    #获取sheet最大行数(指定sheet对象)
    def get_max_row1(self, sh):
        #初始化
        try:
            max_row = sh.max_row
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return max_row


    #获取sheet最小行数(指定sheet对象)
    def get_min_row1(self, sh):
        #初始化
        try:
            min_row = sh.min_row
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return min_row


    #获取sheet最大列数(指定sheet对象)
    def get_max_column1(self, sh):
        #初始化
        try:
            max_column = sh.max_column
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return max_column


    #获取sheet最小列数(指定sheet对象)
    def get_min_column1(self, sh):
        #初始化
        try:
            min_column = sh.min_column
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return min_column



    #读单元格值(指定位置，例：A4)
    def readCell1(self, sh, position):
        #初始化
        value = None
        try:
            value = sh[position].value
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return value



    #读单元格值(指定行、列。行从1开始；列从1开始)
    def readCell2(self, sh, row, column):
        #初始化
        value = None
        try:
            value = sh.cell(row = row, column = column).value
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return value



    #写单元格值(指定位置，例：A4)
    def writeCell1(self, sh, position, value):
        #初始化
        try:
            sh[position] = value
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass



    #写单元格值(指定行、列)
    def writeCell2(self, sh, row1, column1, value1):
        #初始化
        try:
            sh.cell(row = row1, column = column1, value = value1)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass



    #保存excel
    def saveExcel(self, wb, fileName):
        try:
            wb.save(fileName)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass


    #######################################################################################
    #excel读数据
    def ExcelUtil_read(self, fileName):
        try:
            #初始化
            list = []    #输出数组
            #获取workbook类对象（已经存在的excel）
            wb = self.getWb1(fileName)
            #获取第一个sheet页对象
            sh = self.getSheet1(wb, 0)
            #获取sheet中最大行数、列数
            max_row =  self.get_max_row1(sh)
            max_column = self.get_max_column1(sh)
            #遍历测试数据，存入二维数组中
            for i in range(2,max_row+1):
                list.append([])
                for j in range(1,max_column+1):
                    value = self.readCell2(sh,i,j)
                    list[i-2].append(value)
            #保存excel
            wb = self.saveExcel(wb, fileName)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return list


    #######################################################################################
    #excel写数据
    def ExcelUtil_write(self):
        try:
            self.writeCell1()
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return ""



    #######################################################################################
    #新建excel文件
    def ExcelUtil_create(self, fileName):
        try:
            print(fileName)
            #获取workbook类对象（新建excel）
            wb = self.createWb1(fileName)
            #新建sheet页
            sh = self.createSheet1(wb, "new1", 0)
           #保存excel
            wb = self.saveExcel(wb, fileName)
        except Exception as e:
            print(e.args)
            print(traceback.format_exc())
        finally:
            pass
        return ""



if __name__ == "__main__":
    try:
        #初始化
        fileName = r"E:\workspace-python\app_testcase.xlsx"   #excel文件路径
        #实例化对象
        excelutil = ExcelUtil(fileName)
        #调用ExcelUtil_read()方法
        list = excelutil.ExcelUtil_read(fileName)
        print(list)
    except Exception as e:
        print(e.args)
        print(traceback.format_exc())
    finally:
        pass